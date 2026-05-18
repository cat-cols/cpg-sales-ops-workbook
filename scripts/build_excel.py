"""
build_excel.py
==============
Wyld Sales Operations — Excel Workbook Builder

Produces wyld_sales_report.xlsx with 8 fully formatted sheets:

  0. Dashboard       KPI cards + charts (the executive view)
  1. Orders          Full order register with filters + conditional formatting
  2. Order Items     Line-item detail, filterable by SKU / market
  3. Revenue         By market, by rep, by format — pivot-ready
  4. Product Performance  SKU velocity, revenue, compliance flag rate
  5. Inventory       Stock levels, reorder alerts, warehouse breakdown
  6. Account Health  Customer AR aging, license status, credit holds
  7. QC Report       Prioritized flag queue with recommended actions

Usage:
    python build_excel.py
"""

import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (Alignment, Border, Font, GradientFill,
                              PatternFill, Side)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.filters import AutoFilter

from database import get_connection

# ---------------------------------------------------------------------------
# Brand colours
# ---------------------------------------------------------------------------
WYLD_GREEN      = "2D6A4F"
WYLD_GREEN_LITE = "D8F3DC"
WYLD_DARK       = "1B1B2F"
WHITE           = "FFFFFF"
LIGHT_GREY      = "F5F5F5"
MID_GREY        = "D9D9D9"
RED             = "C0392B"
AMBER           = "E67E22"
GREEN_OK        = "27AE60"
YELLOW_WARN     = "F9E79F"
RED_FILL        = "FADBD8"
GREEN_FILL      = "D5F5E3"
HEADER_BLUE     = "1A3C5E"

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def hfont(bold=True, size=11, color=WHITE, name="Arial"):
    return Font(bold=bold, size=size, color=color, name=name)

def bfont(bold=False, size=10, color="000000", name="Arial"):
    return Font(bold=bold, size=size, color=color, name=name)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border(style="thin"):
    s = Side(style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)

def money_fmt():
    return '$#,##0.00'

def pct_fmt():
    return '0.0%'

def apply_header_row(ws, row_num, headers, col_widths=None,
                     bg=HEADER_BLUE, fg=WHITE):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.font      = Font(bold=True, size=10, color=fg, name="Arial")
        cell.fill      = fill(bg)
        cell.alignment = center()
        cell.border    = border()
        if col_widths and col <= len(col_widths):
            ws.column_dimensions[get_column_letter(col)].width = col_widths[col-1]

def write_df(ws, df, start_row=2, number_cols=None,
             pct_cols=None, date_cols=None, stripe=True):
    """Write DataFrame rows with alternating row shading."""
    number_cols = number_cols or []
    pct_cols    = pct_cols    or []
    date_cols   = date_cols   or []

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
        excel_row = start_row + r_idx
        row_fill  = fill(LIGHT_GREY) if stripe and r_idx % 2 == 0 else fill(WHITE)
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.fill      = row_fill
            cell.border    = border()
            cell.font      = bfont()
            cell.alignment = left()
            col_name = df.columns[c_idx - 1]
            if col_name in number_cols:
                cell.number_format = money_fmt()
                cell.alignment     = Alignment(horizontal="right", vertical="center")
            elif col_name in pct_cols:
                cell.number_format = '0.0"%"'
                cell.alignment     = Alignment(horizontal="right", vertical="center")
            elif col_name in date_cols:
                cell.alignment = Alignment(horizontal="center", vertical="center")

def freeze_and_filter(ws, freeze="A2", filter_ref=None):
    ws.freeze_panes = freeze
    if filter_ref:
        ws.auto_filter.ref = filter_ref

# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_dashboard(ws, conn):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    # ── Title banner ──────────────────────────────────────────────────────
    ws.merge_cells("B1:N3")
    title = ws["B1"]
    title.value     = "WYLD DISTRIBUTION — SALES OPERATIONS DASHBOARD"
    title.font      = Font(bold=True, size=18, color=WHITE, name="Arial")
    title.fill      = fill(WYLD_GREEN)
    title.alignment = center()

    ws.merge_cells("B4:N4")
    sub = ws["B4"]
    sub.value     = "Reporting period: Jan 2023 – Mar 2025  |  All US Markets + Canada"
    sub.font      = Font(italic=True, size=10, color="555555", name="Arial")
    sub.fill      = fill(WYLD_GREEN_LITE)
    sub.alignment = center()

    for r in range(1, 5):
        ws.row_dimensions[r].height = 20

    # ── KPI Cards (row 6-10) ──────────────────────────────────────────────
    kpis_sql = {
        "Total Revenue":    "SELECT ROUND(SUM(order_total),2) FROM orders WHERE status IN ('Invoiced','Fulfilled')",
        "Total Orders":     "SELECT COUNT(*) FROM orders",
        "Invoiced Orders":  "SELECT COUNT(*) FROM orders WHERE status='Invoiced'",
        "Avg Order Value":  "SELECT ROUND(AVG(order_total),2) FROM orders WHERE status IN ('Invoiced','Fulfilled')",
        "Compliance Flags": "SELECT COUNT(*) FROM order_items WHERE thc_compliance_flag=1",
        "Open Orders":      "SELECT COUNT(*) FROM orders WHERE status IN ('Pending','Processing')",
    }
    kpi_values = {}
    for label, sql in kpis_sql.items():
        val = conn.execute(sql).fetchone()[0] or 0
        kpi_values[label] = val

    kpi_cards = [
        ("Total Revenue",    f"${kpi_values['Total Revenue']:,.0f}",       WYLD_GREEN,  WHITE),
        ("Total Orders",     f"{kpi_values['Total Orders']:,}",             HEADER_BLUE, WHITE),
        ("Invoiced Orders",  f"{kpi_values['Invoiced Orders']:,}",          "27AE60",    WHITE),
        ("Avg Order Value",  f"${kpi_values['Avg Order Value']:,.2f}",      "1A5276",    WHITE),
        ("Compliance Flags", f"{kpi_values['Compliance Flags']:,}",         "922B21",    WHITE),
        ("Open Orders",      f"{kpi_values['Open Orders']:,}",              "784212",    WHITE),
    ]

    ws.row_dimensions[6].height = 14
    ws.row_dimensions[7].height = 30
    ws.row_dimensions[8].height = 22
    ws.row_dimensions[9].height = 10

    card_cols = [("B", "C"), ("D", "E"), ("F", "G"),
                 ("H", "I"), ("J", "K"), ("L", "M")]
    for (start_col, end_col), (label, value, bg, fg) in zip(card_cols, kpi_cards):
        ws.merge_cells(f"{start_col}7:{end_col}8")
        label_cell = ws[f"{start_col}6"]
        val_cell   = ws[f"{start_col}7"]

        label_cell.value     = label
        label_cell.font      = Font(bold=True, size=9, color="555555", name="Arial")
        label_cell.alignment = center()
        ws.merge_cells(f"{start_col}6:{end_col}6")

        val_cell.value     = value
        val_cell.font      = Font(bold=True, size=16, color=fg, name="Arial")
        val_cell.fill      = fill(bg)
        val_cell.alignment = center()

        for row in range(6, 10):
            ws.row_dimensions[row].height = 18

    # ── Revenue by Market table (B11:E37) ────────────────────────────────
    ws["B10"] = "Revenue by Market"
    ws["B10"].font = Font(bold=True, size=11, color=WYLD_DARK, name="Arial")
    ws.merge_cells("B10:E10")

    mkt_df = pd.read_sql_query("""
        SELECT market,
               total_orders,
               total_revenue,
               invoice_rate_pct
        FROM vw_revenue_by_market
        WHERE total_revenue > 0
        ORDER BY total_revenue DESC
        LIMIT 24
    """, conn)
    mkt_headers = ["Market", "Orders", "Revenue ($)", "Invoice Rate (%)"]
    apply_header_row(ws, 11, mkt_headers,
                     col_widths=[10, 9, 14, 16], bg=WYLD_GREEN)
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = [10, 9, 14, 16][["B","C","D","E"].index(col)]
    for r_idx, row in enumerate(mkt_df.itertuples(index=False), 12):
        cells = [row.market, row.total_orders,
                 row.total_revenue, row.invoice_rate_pct]
        bg_color = LIGHT_GREY if r_idx % 2 == 0 else WHITE
        for c_idx, val in enumerate(cells, 2):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill   = fill(bg_color)
            cell.font   = bfont(size=9)
            cell.border = border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c_idx == 4:
                cell.number_format = money_fmt()
            elif c_idx == 5:
                cell.number_format = '0.0"%"'

    # ── Rep Performance table (G10:M16) ──────────────────────────────────
    ws["G10"] = "Rep Performance"
    ws["G10"].font = Font(bold=True, size=11, color=WYLD_DARK, name="Arial")
    ws.merge_cells("G10:M10")

    rep_df = pd.read_sql_query("""
        SELECT rep_name, region, total_orders,
               total_revenue, avg_order_value, invoice_rate_pct
        FROM vw_revenue_by_rep
        ORDER BY total_revenue DESC
    """, conn)
    rep_headers = ["Rep Name", "Region", "Orders", "Revenue ($)", "Avg Order ($)", "Invoice Rate (%)"]
    apply_header_row(ws, 11, rep_headers, bg=HEADER_BLUE)
    col_positions = list(range(7, 13))
    widths = [20, 14, 9, 14, 14, 16]
    for i, (col_num, w) in enumerate(zip(col_positions, widths)):
        ws.column_dimensions[get_column_letter(col_num)].width = w
        ws.cell(row=11, column=col_num).value    = rep_headers[i]
        ws.cell(row=11, column=col_num).font     = Font(bold=True, size=10, color=WHITE, name="Arial")
        ws.cell(row=11, column=col_num).fill     = fill(HEADER_BLUE)
        ws.cell(row=11, column=col_num).border   = border()
        ws.cell(row=11, column=col_num).alignment = center()

    for r_idx, row in enumerate(rep_df.itertuples(index=False), 12):
        bg_color = LIGHT_GREY if r_idx % 2 == 0 else WHITE
        vals = [row.rep_name, row.region, row.total_orders,
                row.total_revenue, row.avg_order_value, row.invoice_rate_pct]
        for c_idx, (col_num, val) in enumerate(zip(col_positions, vals)):
            cell = ws.cell(row=r_idx, column=col_num, value=val)
            cell.fill      = fill(bg_color)
            cell.font      = bfont(size=9)
            cell.border    = border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c_idx in (3, 4):
                cell.number_format = money_fmt()
            elif c_idx == 5:
                cell.number_format = '0.0"%"'

    # ── Order Status Breakdown table (G20:J27) ────────────────────────────
    ws["G20"] = "Order Status Breakdown"
    ws["G20"].font = Font(bold=True, size=11, color=WYLD_DARK, name="Arial")
    ws.merge_cells("G20:J20")

    status_df = pd.read_sql_query("""
        SELECT status,
               COUNT(*) AS orders,
               ROUND(SUM(order_total),2) AS total_value,
               ROUND(100.0*COUNT(*)/1200.0, 1) AS pct
        FROM orders
        GROUP BY status ORDER BY orders DESC
    """, conn)
    status_headers = ["Status", "Orders", "Value ($)", "% of Total"]
    apply_header_row(ws, 21, status_headers, bg=WYLD_GREEN)
    for i, h in enumerate(status_headers):
        col_num = 7 + i
        ws.cell(row=21, column=col_num).value     = h
        ws.cell(row=21, column=col_num).font      = Font(bold=True, size=10, color=WHITE, name="Arial")
        ws.cell(row=21, column=col_num).fill      = fill(WYLD_GREEN)
        ws.cell(row=21, column=col_num).border    = border()
        ws.cell(row=21, column=col_num).alignment = center()

    STATUS_COLORS = {
        "Invoiced": GREEN_FILL, "Fulfilled": "D6EAF8",
        "Rejected": RED_FILL,   "Returned":  YELLOW_WARN,
        "Refusal":  YELLOW_WARN,"Pending":   "EBF5FB",
        "Processing": "EAF2FF",
    }
    for r_idx, row in enumerate(status_df.itertuples(index=False), 22):
        row_bg = STATUS_COLORS.get(row.status, WHITE)
        vals   = [row.status, row.orders, row.total_value, row.pct]
        for i, val in enumerate(vals):
            col_num = 7 + i
            cell = ws.cell(row=r_idx, column=col_num, value=val)
            cell.fill      = fill(row_bg)
            cell.font      = bfont(size=9)
            cell.border    = border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if i == 2:
                cell.number_format = money_fmt()
            elif i == 3:
                cell.number_format = '0.0"%"'

    # ── Monthly Revenue trend table (B38:E50) ─────────────────────────────
    ws["B38"] = "Monthly Revenue Trend (2024)"
    ws["B38"].font = Font(bold=True, size=11, color=WYLD_DARK, name="Arial")
    ws.merge_cells("B38:E38")

    monthly_df = pd.read_sql_query("""
        SELECT SUBSTR(order_date,1,7) AS month,
               COUNT(*) AS orders,
               ROUND(SUM(order_total),2) AS revenue
        FROM orders
        WHERE status IN ('Invoiced','Fulfilled')
          AND order_date >= '2024-01-01'
          AND order_date < '2025-01-01'
        GROUP BY month ORDER BY month
    """, conn)
    apply_header_row(ws, 39, ["Month", "Orders", "Revenue ($)", ""], bg=WYLD_GREEN)
    for i, h in enumerate(["Month", "Orders", "Revenue ($)"]):
        ws.cell(row=39, column=2+i).value     = h
        ws.cell(row=39, column=2+i).font      = Font(bold=True, size=10, color=WHITE, name="Arial")
        ws.cell(row=39, column=2+i).fill      = fill(WYLD_GREEN)
        ws.cell(row=39, column=2+i).border    = border()
        ws.cell(row=39, column=2+i).alignment = center()
    for r_idx, row in enumerate(monthly_df.itertuples(index=False), 40):
        bg_color = LIGHT_GREY if r_idx % 2 == 0 else WHITE
        for i, val in enumerate([row.month, row.orders, row.revenue]):
            cell = ws.cell(row=r_idx, column=2+i, value=val)
            cell.fill = fill(bg_color); cell.font = bfont(size=9)
            cell.border = border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if i == 2:
                cell.number_format = money_fmt()

    # ── Bar chart: Revenue by Market ──────────────────────────────────────
    chart1 = BarChart()
    chart1.type    = "col"
    chart1.title   = "Revenue by Market (Top 12)"
    chart1.y_axis.title = "Revenue ($)"
    chart1.x_axis.title = "Market"
    chart1.style   = 10
    chart1.width   = 18
    chart1.height  = 12

    data_ref = Reference(ws, min_col=4, min_row=11,
                          max_row=11 + min(11, len(mkt_df)))
    cats_ref = Reference(ws, min_col=2, min_row=12,
                          max_row=11 + min(11, len(mkt_df)))
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.series[0].graphicalProperties.solidFill = WYLD_GREEN
    ws.add_chart(chart1, "G30")

    # ── Line chart: Monthly Revenue ───────────────────────────────────────
    chart2 = LineChart()
    chart2.title  = "Monthly Revenue 2024"
    chart2.y_axis.title = "Revenue ($)"
    chart2.x_axis.title = "Month"
    chart2.style  = 10
    chart2.width  = 18
    chart2.height = 12

    n_months = len(monthly_df)
    rev_ref  = Reference(ws, min_col=4, min_row=39, max_row=39 + n_months)
    mon_ref  = Reference(ws, min_col=2, min_row=40, max_row=39 + n_months)
    chart2.add_data(rev_ref, titles_from_data=True)
    chart2.set_categories(mon_ref)
    chart2.series[0].graphicalProperties.line.solidFill = WYLD_GREEN
    chart2.series[0].graphicalProperties.line.width = 25000
    ws.add_chart(chart2, "B52")


def build_orders_sheet(ws, conn):
    ws.sheet_view.showGridLines = False

    df = pd.read_sql_query("""
        SELECT o.order_id, o.order_date, o.status, o.market,
               o.account_name, c.tier, o.rep_id,
               r.rep_name, o.distribution_channel,
               o.payment_terms, o.po_number,
               o.order_total, o.has_compliance_flag,
               o.rejection_reason, o.fulfillment_date, o.invoice_date,
               o.notes
        FROM orders o
        JOIN customers c ON o.customer_id = c.customer_id
        JOIN sales_reps r ON o.rep_id = r.rep_id
        ORDER BY o.order_date DESC
    """, conn)

    headers = ["Order ID", "Order Date", "Status", "Market", "Account Name",
               "Tier", "Rep ID", "Rep Name", "Channel", "Payment Terms",
               "PO Number", "Order Total", "Compliance Flag",
               "Rejection Reason", "Fulfillment Date", "Invoice Date", "Notes"]
    widths  = [13, 13, 13, 9, 28, 11, 10, 20, 18, 14, 14, 14, 16, 22, 16, 13, 22]

    # Title
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    title = ws["A1"]
    title.value     = "Order Register — Wyld Distribution Co."
    title.font      = Font(bold=True, size=13, color=WHITE, name="Arial")
    title.fill      = fill(WYLD_GREEN)
    title.alignment = center()
    ws.row_dimensions[1].height = 24

    apply_header_row(ws, 2, headers, col_widths=widths, bg=HEADER_BLUE)

    display_df = df.copy()
    display_df.columns = headers
    write_df(ws, display_df, start_row=3,
             number_cols=["Order Total"],
             date_cols=["Order Date", "Fulfillment Date", "Invoice Date"])

    # Conditional formatting: status column (col 3)
    last_row = 2 + len(df)
    status_col = "C"
    STATUS_FORMATS = {
        "Invoiced":   GREEN_FILL,
        "Fulfilled":  "D6EAF8",
        "Rejected":   RED_FILL,
        "Returned":   YELLOW_WARN,
        "Refusal":    YELLOW_WARN,
        "Pending":    "EBF5FB",
        "Processing": "EAF2FF",
    }
    for status, hex_col in STATUS_FORMATS.items():
        ws.conditional_formatting.add(
            f"{status_col}3:{status_col}{last_row}",
            FormulaRule(
                formula=[f'C3="{status}"'],
                fill=fill(hex_col),
                font=bfont(bold=True)
            )
        )

    # Red flag compliance column (col 13 = M now with 17 cols)
    ws.conditional_formatting.add(
        f"M3:M{last_row}",
        CellIsRule(operator="equal", formula=["1"],
                   fill=fill(RED_FILL),
                   font=Font(bold=True, color=RED, name="Arial"))
    )

    freeze_and_filter(ws, freeze="A3",
                      filter_ref=f"A2:{get_column_letter(len(headers))}2")


def build_order_items_sheet(ws, conn):
    ws.sheet_view.showGridLines = False

    df = pd.read_sql_query("""
        SELECT oi.order_id, o.order_date, o.market, o.status,
               oi.sku, oi.product_name, oi.format,
               oi.qty_cases, oi.unit_price, oi.line_total,
               oi.thc_per_unit_mg, oi.thc_limit_market,
               oi.thc_compliance_flag, oi.cold_chain_required
        FROM order_items oi
        JOIN orders o ON oi.order_id = o.order_id
        ORDER BY o.order_date DESC, oi.order_id
    """, conn)

    headers = ["Order ID", "Order Date", "Market", "Status", "SKU",
               "Product Name", "Format", "Qty (Cases)", "Unit Price ($)",
               "Line Total ($)", "THC/Unit (mg)", "Market Limit (mg)",
               "THC Flag", "Cold Chain"]
    widths  = [13, 12, 9, 12, 10, 30, 11, 12, 14, 14, 13, 15, 10, 11]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    title = ws["A1"]
    title.value     = "Order Line Items — Wyld Distribution Co."
    title.font      = Font(bold=True, size=13, color=WHITE, name="Arial")
    title.fill      = fill(WYLD_GREEN)
    title.alignment = center()
    ws.row_dimensions[1].height = 24

    apply_header_row(ws, 2, headers, col_widths=widths, bg=HEADER_BLUE)

    display_df = df.copy()
    display_df.columns = headers
    write_df(ws, display_df, start_row=3,
             number_cols=["Unit Price ($)", "Line Total ($)"])

    last_row = 2 + len(df)
    ws.conditional_formatting.add(
        f"M3:M{last_row}",
        CellIsRule(operator="equal", formula=["1"],
                   fill=fill(RED_FILL),
                   font=Font(bold=True, color=RED, name="Arial"))
    )
    ws.conditional_formatting.add(
        f"J3:J{last_row}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max",   end_color="2D6A4F")
    )

    freeze_and_filter(ws, freeze="A3",
                      filter_ref=f"A2:{get_column_letter(len(headers))}2")


def build_revenue_sheet(ws, conn):
    ws.sheet_view.showGridLines = False

    # Section 1: By Market
    ws.merge_cells("A1:H1")
    ws["A1"].value     = "Revenue Analysis — Wyld Distribution Co."
    ws["A1"].font      = Font(bold=True, size=13, color=WHITE, name="Arial")
    ws["A1"].fill      = fill(WYLD_GREEN)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 24

    ws["A3"] = "Revenue by Market"
    ws["A3"].font = Font(bold=True, size=11, color=WYLD_DARK, name="Arial")

    mkt_df = pd.read_sql_query("""
        SELECT market, is_canada, total_orders, invoiced_orders,
               rejected_orders, returned_orders, total_revenue,
               avg_order_value, invoice_rate_pct, rejection_rate_pct
        FROM vw_revenue_by_market
        ORDER BY total_revenue DESC
    """, conn)
    mkt_headers = ["Market", "Canada?", "Total Orders", "Invoiced",
                   "Rejected", "Returned", "Total Revenue ($)",
                   "Avg Order ($)", "Invoice Rate (%)", "Rejection Rate (%)"]
    widths = [10, 9, 13, 11, 10, 10, 16, 14, 15, 16]
    apply_header_row(ws, 4, mkt_headers, col_widths=widths, bg=HEADER_BLUE)
    write_df(ws, mkt_df, start_row=5,
             number_cols=["total_revenue", "avg_order_value"],
             pct_cols=["invoice_rate_pct", "rejection_rate_pct"])

    # Color scale on revenue column (G)
    last = 4 + len(mkt_df)
    ws.conditional_formatting.add(
        f"G5:G{last}",
        ColorScaleRule(start_type="min", start_color=WYLD_GREEN_LITE,
                       end_type="max",   end_color=WYLD_GREEN)
    )

    # Section 2: By Rep
    offset = last + 3
    ws.cell(row=offset, column=1).value = "Revenue by Sales Rep"
    ws.cell(row=offset, column=1).font  = Font(bold=True, size=11,
                                               color=WYLD_DARK, name="Arial")

    rep_df = pd.read_sql_query("""
        SELECT rep_name, region, total_orders, invoiced_orders,
               rejected_orders, total_revenue, avg_order_value,
               unique_accounts, invoice_rate_pct
        FROM vw_revenue_by_rep ORDER BY total_revenue DESC
    """, conn)
    rep_headers = ["Rep Name", "Region", "Total Orders", "Invoiced",
                   "Rejected", "Revenue ($)", "Avg Order ($)",
                   "Accounts", "Invoice Rate (%)"]
    apply_header_row(ws, offset+1, rep_headers, bg=HEADER_BLUE)
    write_df(ws, rep_df, start_row=offset+2,
             number_cols=["total_revenue", "avg_order_value"],
             pct_cols=["invoice_rate_pct"])

    # Section 3: By Format
    offset2 = offset + len(rep_df) + 4
    ws.cell(row=offset2, column=1).value = "Revenue by Product Format"
    ws.cell(row=offset2, column=1).font  = Font(bold=True, size=11,
                                                color=WYLD_DARK, name="Arial")
    fmt_df = pd.read_sql_query("""
        SELECT p.format,
               COUNT(DISTINCT oi.order_id) AS orders,
               SUM(oi.qty_cases)           AS total_cases,
               ROUND(SUM(oi.line_total),2) AS revenue,
               ROUND(AVG(oi.unit_price),2) AS avg_price
        FROM order_items oi
        JOIN products p ON oi.sku = p.sku
        GROUP BY p.format ORDER BY revenue DESC
    """, conn)
    fmt_headers = ["Format", "Orders", "Cases Sold", "Revenue ($)", "Avg Price ($)"]
    apply_header_row(ws, offset2+1, fmt_headers, bg=HEADER_BLUE)
    write_df(ws, fmt_df, start_row=offset2+2,
             number_cols=["revenue", "avg_price"])

    freeze_and_filter(ws, freeze="A5",
                      filter_ref=f"A4:{get_column_letter(len(mkt_headers))}4")


def build_product_sheet(ws, conn):
    ws.sheet_view.showGridLines = False

    df = pd.read_sql_query("""
        SELECT pp.sku, pp.product_name, pp.format,
               pp.orders_containing_sku, pp.total_cases_sold,
               pp.total_revenue, pp.avg_unit_price,
               pp.compliance_flag_count, pp.flag_rate_pct,
               iv.on_hand_units, iv.available_units,
               iv.reorder_point, iv.stock_status, iv.warehouse_location
        FROM vw_product_performance pp
        JOIN vw_inventory_status iv ON pp.sku = iv.sku
        ORDER BY pp.total_revenue DESC
    """, conn)

    headers = ["SKU", "Product Name", "Format", "Orders", "Cases Sold",
               "Revenue ($)", "Avg Price ($)", "Compliance Flags",
               "Flag Rate (%)", "On Hand", "Available",
               "Reorder Pt", "Stock Status", "Warehouse"]
    widths  = [10, 32, 10, 9, 12, 14, 13, 16, 12, 10, 10, 10, 14, 12]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"].value     = "Product Performance — Wyld Distribution Co."
    ws["A1"].font      = Font(bold=True, size=13, color=WHITE, name="Arial")
    ws["A1"].fill      = fill(WYLD_GREEN)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 24

    apply_header_row(ws, 2, headers, col_widths=widths, bg=HEADER_BLUE)
    display_df = df.copy()
    display_df.columns = headers
    write_df(ws, display_df, start_row=3,
             number_cols=["Revenue ($)", "Avg Price ($)"],
             pct_cols=["Flag Rate (%)"])

    last_row = 2 + len(df)
    # Stock status color coding
    for r in range(3, last_row + 1):
        cell = ws.cell(row=r, column=13)
        if cell.value == "OUT OF STOCK":
            cell.fill = fill(RED_FILL)
            cell.font = Font(bold=True, color=RED, name="Arial", size=10)
        elif cell.value == "REORDER NOW":
            cell.fill = fill(YELLOW_WARN)
            cell.font = Font(bold=True, color=AMBER, name="Arial", size=10)
        elif cell.value == "LOW STOCK":
            cell.fill = fill("FAD7A0")
            cell.font = bfont(bold=True)

    # Color scale on revenue
    ws.conditional_formatting.add(
        f"F3:F{last_row}",
        ColorScaleRule(start_type="min", start_color=WYLD_GREEN_LITE,
                       end_type="max",   end_color=WYLD_GREEN)
    )

    freeze_and_filter(ws, freeze="A3",
                      filter_ref=f"A2:{get_column_letter(len(headers))}2")


def build_inventory_sheet(ws, conn):
    ws.sheet_view.showGridLines = False

    df = pd.read_sql_query("""
        SELECT sku, product_name, format, warehouse_location,
               on_hand_units, allocated_units, available_units,
               reorder_point, below_reorder, stock_status,
               inventory_value, last_received_date
        FROM vw_inventory_status
        ORDER BY available_units ASC
    """, conn)

    headers = ["SKU", "Product Name", "Format", "Warehouse",
               "On Hand", "Allocated", "Available",
               "Reorder Pt", "Below Reorder?", "Stock Status",
               "Inventory Value ($)", "Last Received"]
    widths  = [10, 32, 10, 12, 10, 11, 10, 10, 14, 14, 18, 14]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"].value     = "Inventory Status — Wyld Distribution Co."
    ws["A1"].font      = Font(bold=True, size=13, color=WHITE, name="Arial")
    ws["A1"].fill      = fill(WYLD_GREEN)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 24

    apply_header_row(ws, 2, headers, col_widths=widths, bg=HEADER_BLUE)
    display_df = df.copy()
    display_df.columns = headers
    write_df(ws, display_df, start_row=3,
             number_cols=["Inventory Value ($)"])

    last_row = 2 + len(df)
    for r in range(3, last_row + 1):
        status_cell = ws.cell(row=r, column=10)
        if status_cell.value == "OUT OF STOCK":
            for c in range(1, len(headers)+1):
                ws.cell(row=r, column=c).fill = fill(RED_FILL)
            status_cell.font = Font(bold=True, color=RED, name="Arial", size=10)
        elif status_cell.value == "REORDER NOW":
            status_cell.fill = fill(YELLOW_WARN)
            status_cell.font = Font(bold=True, color=AMBER, name="Arial", size=10)

    ws.conditional_formatting.add(
        f"G3:G{last_row}",
        ColorScaleRule(start_type="min", start_color="FADBD8",
                       end_type="max",   end_color=GREEN_FILL)
    )

    freeze_and_filter(ws, freeze="A3",
                      filter_ref=f"A2:{get_column_letter(len(headers))}2")


def build_account_health_sheet(ws, conn):
    ws.sheet_view.showGridLines = False

    df = pd.read_sql_query("""
        SELECT customer_id, account_name, tier, market, customer_type,
               payment_terms, credit_limit, on_credit_hold,
               license_expiry, license_status, lifetime_orders,
               lifetime_revenue, avg_order_value, last_order_date,
               rejected_orders, returned_orders, open_ar_amount
        FROM vw_account_health
        ORDER BY lifetime_revenue DESC
    """, conn)

    headers = ["Customer ID", "Account Name", "Tier", "Market", "Type",
               "Payment Terms", "Credit Limit ($)", "On Hold?",
               "License Expiry", "License Status", "Lifetime Orders",
               "Lifetime Revenue ($)", "Avg Order ($)", "Last Order",
               "Rejected", "Returned", "Open AR ($)"]
    widths  = [12, 28, 10, 9, 16, 14, 15, 10, 14, 15, 14, 18, 13, 13, 10, 10, 13]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"].value     = "Account Health — Wyld Distribution Co."
    ws["A1"].font      = Font(bold=True, size=13, color=WHITE, name="Arial")
    ws["A1"].fill      = fill(WYLD_GREEN)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 24

    apply_header_row(ws, 2, headers, col_widths=widths, bg=HEADER_BLUE)
    display_df = df.copy()
    display_df.columns = headers
    write_df(ws, display_df, start_row=3,
             number_cols=["Credit Limit ($)", "Lifetime Revenue ($)",
                          "Avg Order ($)", "Open AR ($)"],
             date_cols=["License Expiry", "Last Order"])

    last_row = 2 + len(df)
    # Credit hold highlight
    ws.conditional_formatting.add(
        f"H3:H{last_row}",
        CellIsRule(operator="equal", formula=["1"],
                   fill=fill(RED_FILL),
                   font=Font(bold=True, color=RED, name="Arial"))
    )
    # License status
    ws.conditional_formatting.add(
        f"J3:J{last_row}",
        FormulaRule(formula=['J3="EXPIRED"'],
                    fill=fill(RED_FILL),
                    font=Font(bold=True, color=RED, name="Arial"))
    )
    ws.conditional_formatting.add(
        f"J3:J{last_row}",
        FormulaRule(formula=['J3="EXPIRING SOON"'],
                    fill=fill(YELLOW_WARN),
                    font=Font(bold=True, color=AMBER, name="Arial"))
    )
    # Tier color
    TIER_COLORS = {"Platinum": "D7BDE2", "Gold": "FAD7A0",
                   "Silver":   "D5DBDB", "Bronze": "E59866"}
    for tier, hex_color in TIER_COLORS.items():
        ws.conditional_formatting.add(
            f"C3:C{last_row}",
            FormulaRule(formula=[f'C3="{tier}"'], fill=fill(hex_color))
        )

    freeze_and_filter(ws, freeze="A3",
                      filter_ref=f"A2:{get_column_letter(len(headers))}2")


def build_qc_sheet(ws, conn):
    ws.sheet_view.showGridLines = False

    qc_path = Path("data/qc_report.csv")
    if not qc_path.exists():
        ws["A1"] = "QC report not found — run quality_check.py first"
        return

    df = pd.read_csv(qc_path)
    # Sort: CRITICAL first, then WARNING, then INFO
    sev_order = {"CRITICAL": 0, "WARNING": 1, "INFO": 2}
    df["_sort"] = df["severity"].map(sev_order)
    df = df.sort_values("_sort").drop(columns="_sort")

    headers = ["Severity", "Category", "Check Name", "Table",
               "Record ID", "Field", "Value", "Detail", "Recommended Action"]
    widths  = [11, 18, 28, 16, 14, 16, 22, 45, 40]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"].value     = "Data Quality & Compliance Report — Wyld Distribution Co."
    ws["A1"].font      = Font(bold=True, size=13, color=WHITE, name="Arial")
    ws["A1"].fill      = fill("7B241C")
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 24

    # Summary banner row
    n_crit = (df["severity"] == "CRITICAL").sum()
    n_warn = (df["severity"] == "WARNING").sum()
    n_info = (df["severity"] == "INFO").sum()
    ws.merge_cells(f"A2:{get_column_letter(len(headers))}2")
    ws["A2"].value = (f"  Total Flags: {len(df)}   |   "
                      f"● CRITICAL: {n_crit}   "
                      f"▲ WARNING: {n_warn}   "
                      f"ℹ INFO: {n_info}")
    ws["A2"].font      = Font(bold=True, size=10, color=WHITE, name="Arial")
    ws["A2"].fill      = fill("922B21")
    ws["A2"].alignment = left()
    ws.row_dimensions[2].height = 18

    apply_header_row(ws, 3, headers, col_widths=widths, bg="7B241C", fg=WHITE)
    display_df = df[["severity","category","check_name","table",
                      "record_id","field","value","detail",
                      "recommended_action"]].copy()
    write_df(ws, display_df, start_row=4, stripe=False)

    last_row = 3 + len(df)
    SEV_STYLES = {
        "CRITICAL": (RED_FILL,    RED,   True),
        "WARNING":  (YELLOW_WARN, AMBER, True),
        "INFO":     ("EBF5FB",    "1A5276", False),
    }
    for r in range(4, last_row + 1):
        sev = ws.cell(row=r, column=1).value
        if sev in SEV_STYLES:
            bg, fc, bold = SEV_STYLES[sev]
            ws.cell(row=r, column=1).fill = fill(bg)
            ws.cell(row=r, column=1).font = Font(bold=bold, color=fc,
                                                  size=10, name="Arial")
        # Wrap detail + action columns
        for c in (8, 9):
            ws.cell(row=r, column=c).alignment = Alignment(
                wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 28

    freeze_and_filter(ws, freeze="A4",
                      filter_ref=f"A3:{get_column_letter(len(headers))}3")


# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------

SHEETS = [
    ("Dashboard",           build_dashboard),
    ("Orders",              build_orders_sheet),
    ("Order Items",         build_order_items_sheet),
    ("Revenue",             build_revenue_sheet),
    ("Product Performance", build_product_sheet),
    ("Inventory",           build_inventory_sheet),
    ("Account Health",      build_account_health_sheet),
    ("QC Report",           build_qc_sheet),
]

TAB_COLORS = [
    WYLD_GREEN, HEADER_BLUE, "1A5276", "117A65",
    "6C3483",   "784212",    "1B2631", "7B241C",
]


def build_workbook() -> Path:
    out_path = Path("wyld_sales_report.xlsx")
    conn     = get_connection()
    wb       = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    print("\n Wyld Sales Ops — Excel Workbook Builder")
    print("=" * 42)

    for (name, builder), tab_color in zip(SHEETS, TAB_COLORS):
        print(f"  Building sheet: {name}...")
        ws = wb.create_sheet(title=name)
        ws.sheet_properties.tabColor = tab_color
        builder(ws, conn)

    conn.close()
    wb.save(out_path)
    print(f"\n  Saved → {out_path}")
    return out_path


if __name__ == "__main__":
    out = build_workbook()
    print(f"  Open: {out.resolve()}")
