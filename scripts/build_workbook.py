from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.styles.numbers import FORMAT_DATE_DATETIME
import datetime

wb = Workbook()

# ── Palette ──────────────────────────────────────────────────────────────────
WYLD_GREEN   = "2D6A4F"
WYLD_GREEN2  = "40916C"
WYLD_LIGHT   = "D8F3DC"
WYLD_ACCENT  = "B7E4C7"
WHITE        = "FFFFFF"
LIGHT_GRAY   = "F5F5F5"
MID_GRAY     = "D9D9D9"
DARK_GRAY    = "595959"
HEADER_FONT  = Font(name="Arial", bold=True, color=WHITE, size=10)
BODY_FONT    = Font(name="Arial", size=10)
TITLE_FONT   = Font(name="Arial", bold=True, size=14, color=WYLD_GREEN)
SUB_FONT     = Font(name="Arial", bold=True, size=10, color=DARK_GRAY)
LABEL_FONT   = Font(name="Arial", bold=True, size=10)
GREEN_FILL   = PatternFill("solid", fgColor=WYLD_GREEN)
GREEN2_FILL  = PatternFill("solid", fgColor=WYLD_GREEN2)
LIGHT_FILL   = PatternFill("solid", fgColor=WYLD_LIGHT)
ACCENT_FILL  = PatternFill("solid", fgColor=WYLD_ACCENT)
GRAY_FILL    = PatternFill("solid", fgColor=LIGHT_GRAY)
MID_FILL     = PatternFill("solid", fgColor=MID_GRAY)
WHITE_FILL   = PatternFill("solid", fgColor=WHITE)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT         = Alignment(horizontal="left",   vertical="center")
RIGHT        = Alignment(horizontal="right",  vertical="center")

def thin_border(top=False, bottom=False, left=False, right=False):
    s = Side(style="thin", color="BBBBBB")
    t = s if top    else Side(style=None)
    b = s if bottom else Side(style=None)
    l = s if left   else Side(style=None)
    r = s if right  else Side(style=None)
    return Border(top=t, bottom=b, left=l, right=r)

def full_border():
    s = Side(style="thin", color="BBBBBB")
    return Border(top=s, bottom=s, left=s, right=s)

def header_cell(ws, row, col, value, fill=None, font=None, align=None, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = fill  or GREEN_FILL
    c.font   = font  or HEADER_FONT
    c.alignment = align or CENTER
    c.border = full_border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def style_row(ws, row, cols, fill, font=None, align=None, num_fmt=None, border=True):
    for col in cols:
        c = ws.cell(row=row, column=col)
        c.fill = fill
        if font:  c.font = font
        if align: c.alignment = align
        if num_fmt: c.number_format = num_fmt
        if border: c.border = full_border()


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — PRODUCT CATALOG
# ═══════════════════════════════════════════════════════════════════════════════
ws_prod = wb.active
ws_prod.title = "Product Catalog"
ws_prod.sheet_view.showGridLines = False
ws_prod.freeze_panes = "A3"

ws_prod.row_dimensions[1].height = 36
ws_prod.row_dimensions[2].height = 22

ws_prod.merge_cells("A1:K1")
t = ws_prod["A1"]
t.value = "MYLD — Product Catalog"
t.font  = TITLE_FONT
t.alignment = LEFT
t.fill = WHITE_FILL

prod_headers = ["SKU","Product Name","Category","Unit Price","Case Qty",
                "Case Price","OR","WA","CA","CO","NV","MI","IL","NJ","NY","MA"]
col_widths   = [22, 38, 12, 12, 10, 12, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5]
for i, (h, w) in enumerate(zip(prod_headers, col_widths), 1):
    header_cell(ws_prod, 2, i, h, width=w)

products = [
    ("MLD-GUM-HUCK-10",  "Myld Huckleberry Gummies 10mg THC",        "Edible",   18.00, 12, "✓","✓","✓","✓","✓","","","","",""),
    ("MLD-GUM-RASP-10",  "Myld Raspberry Gummies 10mg THC",           "Edible",   18.00, 12, "✓","✓","✓","✓","✓","✓","✓","","",""),
    ("MLD-GUM-BBLM-1-1", "Myld Blackberry Gummies 1:1 CBD:THC",       "Edible",   20.00, 12, "✓","✓","✓","✓","","","","","",""),
    ("MLD-BEV-HUCK-5",   "Myld Huckleberry Sparkling Water 5mg",      "Beverage",  6.00, 24, "✓","✓","✓","","","","","","",""),
    ("MLD-BEV-LEMON-5",  "Myld Lemon Sparkling Water 5mg",            "Beverage",  6.00, 24, "✓","✓","✓","✓","✓","","","","",""),
    ("MLD-GUM-PEACH-2-1","Myld Peach Gummies 2:1 CBD:THC",            "Edible",   22.00, 12, "✓","","✓","✓","","","","","",""),
    ("MLD-GUM-STRAW-10", "Myld Strawberry Gummies 10mg THC",          "Edible",   18.00, 12, "","","","","","✓","✓","✓","✓","✓"),
]

for r, p in enumerate(products, 3):
    ws_prod.row_dimensions[r].height = 18
    row_fill = GRAY_FILL if r % 2 == 0 else WHITE_FILL
    vals = list(p)
    for c, v in enumerate(vals[:5], 1):
        cell = ws_prod.cell(row=r, column=c, value=v)
        cell.fill   = row_fill
        cell.font   = BODY_FONT
        cell.border = full_border()
        cell.alignment = CENTER if c in (1,3,5) else LEFT
        if c == 4: cell.number_format = '$#,##0.00'
    # Case price formula
    cp = ws_prod.cell(row=r, column=6, value=f"=D{r}*E{r}")
    cp.fill = row_fill; cp.font = Font(name="Arial", size=10, color="000000")
    cp.number_format = '$#,##0.00'; cp.border = full_border(); cp.alignment = RIGHT
    # State availability checkboxes
    for c, v in enumerate(vals[5:], 7):
        cell = ws_prod.cell(row=r, column=c, value=v)
        cell.fill = ACCENT_FILL if v == "✓" else row_fill
        cell.font = Font(name="Arial", size=10, bold=(v=="✓"), color=WYLD_GREEN if v=="✓" else DARK_GRAY)
        cell.border = full_border(); cell.alignment = CENTER


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — ACCOUNT MASTER
# ═══════════════════════════════════════════════════════════════════════════════
ws_acc = wb.create_sheet("Account Master")
ws_acc.sheet_view.showGridLines = False
ws_acc.freeze_panes = "A3"
ws_acc.row_dimensions[1].height = 36
ws_acc.row_dimensions[2].height = 22

ws_acc.merge_cells("A1:H1")
t = ws_acc["A1"]
t.value = "MYLD — Account Master"
t.font  = TITLE_FONT; t.alignment = LEFT; t.fill = WHITE_FILL

acc_headers = ["Account ID","Account Name","State","License #","Payment Terms","Active","Credit Limit","Notes"]
acc_widths   = [14, 30, 8, 18, 15, 9, 14, 30]
for i, (h, w) in enumerate(zip(acc_headers, acc_widths), 1):
    header_cell(ws_acc, 2, i, h, width=w)

accounts = [
    ("ACC-001","Green Leaf Dispensary",    "OR","CDS-1234",    "NET30","Yes",25000,""),
    ("ACC-002","Emerald City Cannabis",    "WA","WA-MJ-2891",  "NET15","Yes",20000,""),
    ("ACC-003","The Healing Center",       "CA","CA-LIC-9921", "NET30","Yes",30000,""),
    ("ACC-004","Mile High Dispensary",     "CO","CO-MED-0044", "NET30","Yes",15000,""),
    ("ACC-005","Silver State Cannabis",    "NV","NV-DIS-7731", "NET15","Yes",18000,""),
    ("ACC-006","Garden State Greens",      "NJ","NJ-CRC-2210", "NET30","No", 0,    "Account on hold — license renewal pending"),
    ("ACC-007","Bay Area Botanicals",      "CA","CA-LIC-3347", "NET30","Yes",22000,""),
]

for r, a in enumerate(accounts, 3):
    ws_acc.row_dimensions[r].height = 18
    row_fill = GRAY_FILL if r % 2 == 0 else WHITE_FILL
    for c, v in enumerate(a, 1):
        cell = ws_acc.cell(row=r, column=c, value=v)
        cell.fill   = row_fill
        cell.font   = BODY_FONT
        cell.border = full_border()
        cell.alignment = CENTER if c in (1,3,5,6) else LEFT
        if c == 7: cell.number_format = '$#,##0'
    # Color Active column
    active_cell = ws_acc.cell(row=r, column=6)
    if a[5] == "Yes":
        active_cell.font = Font(name="Arial", size=10, bold=True, color="2D6A4F")
        active_cell.fill = ACCENT_FILL
    else:
        active_cell.font = Font(name="Arial", size=10, bold=True, color="C0392B")
        active_cell.fill = PatternFill("solid", fgColor="FADBD8")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — COMPLIANCE RULES
# ═══════════════════════════════════════════════════════════════════════════════
ws_comp = wb.create_sheet("Compliance Rules")
ws_comp.sheet_view.showGridLines = False
ws_comp.freeze_panes = "A3"
ws_comp.row_dimensions[1].height = 36
ws_comp.row_dimensions[2].height = 22

ws_comp.merge_cells("A1:I1")
t = ws_comp["A1"]
t.value = "MYLD — State Compliance Rules (Reference)"
t.font  = TITLE_FONT; t.alignment = LEFT; t.fill = WHITE_FILL

comp_headers = ["State","Max THC/Unit (mg)","Max THC/Pkg (mg)","Metrc Required",
                "Returns Allowed","Return Window (days)","Refusals Allowed",
                "Excise Tax Rate","Regulatory Notes"]
comp_widths  = [8, 18, 18, 16, 16, 20, 18, 16, 52]
for i, (h, w) in enumerate(zip(comp_headers, comp_widths), 1):
    header_cell(ws_comp, 2, i, h, width=w)

compliance = [
    ("OR", 10, 100, "Yes", "No",  "",   "Yes", 0.17, "OLCC: No returns post-transfer. Refusals accepted at door."),
    ("WA", 10, 100, "Yes", "No",  "",   "Yes", 0.37, "LCB: Returns prohibited. High excise tax state."),
    ("CA", 10, 100, "Yes", "Yes", 30,   "Yes", 0.15, "DCC: Returns within 30 days for damaged/defective only."),
    ("CO", 10, 100, "Yes", "Yes", 14,   "Yes", 0.15, "MED: Returns 14 days. Metrc transfer required all movements."),
    ("NV", 10, 100, "Yes", "No",  "",   "Yes", 0.10, "CCB: No returns. Refusals must be documented on manifest."),
    ("MI", 10, 100, "Yes", "Yes", 30,   "Yes", 0.10, "MRA: Returns allowed damaged/defective. Metrc required."),
    ("IL", 10, 100, "Yes", "No",  "",   "Yes", 0.10, "IDFPR: No returns permitted post-transfer."),
    ("NJ", 10, 100, "Yes", "No",  "",   "Yes", 0.07, "CRC: Returns not currently permitted under NJ regs."),
    ("NY",  5,  50, "Yes", "Yes", 14,   "Yes", 0.09, "OCM: 5mg per-unit cap. Returns 14 days damaged only."),
    ("MA",  5, 100, "Yes", "Yes", 30,   "Yes", 0.107,"CCC: 5mg per-unit cap. Returns within 30 days."),
    ("OH", 10, 100, "Yes", "No",  "",   "Yes", 0.10, "Division of Cannabis Control. No returns currently."),
    ("AZ", 10, 100, "Yes", "Yes", 14,   "Yes", 0.16, "ADHS: Returns 14 days defective only."),
    ("MO", 10, 100, "Yes", "No",  "",   "Yes", 0.06, "DHSS: No returns. Low excise tax state."),
    ("NM", 10, 100, "Yes", "Yes", 30,   "Yes", 0.12, "RLD: Returns within 30 days damaged product."),
    ("MD", 10, 100, "Yes", "No",  "",   "Yes", 0.09, "MCA: No returns post-delivery."),
    ("OK", 10, 100, "Yes", "Yes", 14,   "Yes", 0.07, "OMMA: Returns 14 days defective. Low-tax medical state."),
]

for r, row in enumerate(compliance, 3):
    ws_comp.row_dimensions[r].height = 18
    row_fill = GRAY_FILL if r % 2 == 0 else WHITE_FILL
    for c, v in enumerate(row, 1):
        cell = ws_comp.cell(row=r, column=c, value=v)
        cell.fill = row_fill
        cell.font = BODY_FONT
        cell.border = full_border()
        cell.alignment = CENTER if c in (1,2,3,4,5,6,7) else LEFT
        if c == 8:
            cell.number_format = '0.0%'
    # Color returns column
    ret_cell = ws_comp.cell(row=r, column=5)
    if row[4] == "Yes":
        ret_cell.font = Font(name="Arial", size=10, bold=True, color="2D6A4F")
    else:
        ret_cell.font = Font(name="Arial", size=10, bold=True, color="C0392B")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — ORDER LOG (main working sheet)
# ═══════════════════════════════════════════════════════════════════════════════
ws_ord = wb.create_sheet("Order Log")
ws_ord.sheet_view.showGridLines = False
ws_ord.freeze_panes = "A4"
ws_ord.row_dimensions[1].height = 40
ws_ord.row_dimensions[2].height = 16
ws_ord.row_dimensions[3].height = 22

ws_ord.merge_cells("A1:R1")
t = ws_ord["A1"]
t.value = "MYLD — Sales Order Log"
t.font  = TITLE_FONT; t.alignment = LEFT; t.fill = WHITE_FILL

ws_ord.merge_cells("A2:R2")
sub = ws_ord["A2"]
sub.value = "One row per order line. Update Status and Fulfilled Qty as orders progress through lifecycle."
sub.font  = Font(name="Arial", size=9, italic=True, color=DARK_GRAY)
sub.alignment = LEFT

ord_headers = [
    "Order ID","Order Date","Account ID","Account Name","State",
    "SKU","Product Name","Ordered Qty","Unit Price","Line Total",
    "Fulfilled Qty","Fulfilled Total","Status","Invoice #","Invoice Date",
    "Due Date","Credit Memo #","Notes"
]
ord_widths = [14,13,11,26,7,22,34,12,12,13,13,15,14,14,14,14,14,30]
for i, (h, w) in enumerate(zip(ord_headers, ord_widths), 1):
    header_cell(ws_ord, 3, i, h, width=w)

# Status dropdown validation
status_dv = DataValidation(
    type="list",
    formula1='"Draft,Submitted,Confirmed,Fulfillment,Invoiced,Closed,Cancelled,Rejected"',
    allow_blank=True
)
status_dv.error = "Invalid status. Choose from the dropdown."
status_dv.errorTitle = "Invalid Status"
ws_ord.add_data_validation(status_dv)
status_dv.sqref = "M4:M500"

today = datetime.date.today()
sample_orders = [
    # ord_id, date, acc_id, acc_name, state, sku, prod_name, ord_qty, unit_price, ful_qty, status, inv, inv_date, due, cm, notes
    ("ORD-A1B2C3D4", today - datetime.timedelta(days=30), "ACC-001","Green Leaf Dispensary","OR",
     "MLD-GUM-HUCK-10","Myld Huckleberry Gummies 10mg",10,18.00,10,"Closed",
     "INV-001", today - datetime.timedelta(days=28), today + datetime.timedelta(days=2),"",""),
    ("ORD-A1B2C3D4", today - datetime.timedelta(days=30), "ACC-001","Green Leaf Dispensary","OR",
     "MLD-GUM-RASP-10","Myld Raspberry Gummies 10mg",  5, 18.00,5,"Closed",
     "INV-001", today - datetime.timedelta(days=28), today + datetime.timedelta(days=2),"",""),
    ("ORD-E5F6G7H8", today - datetime.timedelta(days=14), "ACC-003","The Healing Center","CA",
     "MLD-GUM-HUCK-10","Myld Huckleberry Gummies 10mg",8, 18.00,8,"Invoiced",
     "INV-002", today - datetime.timedelta(days=12), today + datetime.timedelta(days=18),"CM-001","Damaged in transit — 2 units. CM issued."),
    ("ORD-E5F6G7H8", today - datetime.timedelta(days=14), "ACC-003","The Healing Center","CA",
     "MLD-BEV-LEMON-5","Myld Lemon Sparkling Water 5mg",12,6.00, 12,"Invoiced",
     "INV-002", today - datetime.timedelta(days=12), today + datetime.timedelta(days=18),"",""),
    ("ORD-I9J0K1L2", today - datetime.timedelta(days=7),  "ACC-002","Emerald City Cannabis","WA",
     "MLD-GUM-RASP-10","Myld Raspberry Gummies 10mg",  6, 18.00,6,"Fulfillment",
     "","","","",""),
    ("ORD-M3N4O5P6", today - datetime.timedelta(days=3),  "ACC-004","Mile High Dispensary","CO",
     "MLD-GUM-BBLM-1-1","Myld Blackberry Gummies 1:1", 4, 20.00,0,"Confirmed",
     "","","","",""),
    ("ORD-M3N4O5P6", today - datetime.timedelta(days=3),  "ACC-004","Mile High Dispensary","CO",
     "MLD-GUM-PEACH-2-1","Myld Peach Gummies 2:1",     6, 22.00,0,"Confirmed",
     "","","","",""),
    ("ORD-Q7R8S9T0", today - datetime.timedelta(days=1),  "ACC-007","Bay Area Botanicals","CA",
     "MLD-GUM-HUCK-10","Myld Huckleberry Gummies 10mg",15,18.00,0,"Submitted",
     "","","","",""),
    ("ORD-U1V2W3X4", today,                               "ACC-005","Silver State Cannabis","NV",
     "MLD-BEV-HUCK-5","Myld Huckleberry Sparkling Water",24,6.00,0,"Draft",
     "","","","",""),
]

STATUS_COLORS = {
    "Draft":       "F0F0F0",
    "Submitted":   "D6EAF8",
    "Confirmed":   "D5F5E3",
    "Fulfillment": "FDEBD0",
    "Invoiced":    "EBF5FB",
    "Closed":      "EAECEE",
    "Cancelled":   "FADBD8",
    "Rejected":    "FADBD8",
}

for r, row in enumerate(sample_orders, 4):
    ws_ord.row_dimensions[r].height = 18
    status = row[10]
    base_fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))

    ord_qty   = row[7]
    unit_px   = row[8]
    ful_qty   = row[9]
    line_col  = 10   # J = Line Total
    ful_col   = 12   # L = Fulfilled Total

    vals = [
        row[0],  # A Order ID
        row[1],  # B Order Date
        row[2],  # C Account ID
        row[3],  # D Account Name
        row[4],  # E State
        row[5],  # F SKU
        row[6],  # G Product Name
        ord_qty, # H Ordered Qty
        unit_px, # I Unit Price
        f"=H{r}*I{r}",  # J Line Total
        ful_qty, # K Fulfilled Qty
        f"=K{r}*I{r}",  # L Fulfilled Total
        status,  # M Status
        row[11], # N Invoice #
        row[12] if row[12] else "",  # O Invoice Date
        row[13] if row[13] else "",  # P Due Date
        row[14], # Q Credit Memo #
        row[15], # R Notes
    ]

    for c, v in enumerate(vals, 1):
        cell = ws_ord.cell(row=r, column=c, value=v)
        cell.fill   = base_fill
        cell.font   = BODY_FONT
        cell.border = full_border()
        cell.alignment = CENTER if c in (1,3,5,8,11,13) else LEFT
        if c in (9,):  cell.number_format = '$#,##0.00'
        if c in (10,12): cell.number_format = '$#,##0.00'; cell.alignment = RIGHT
        if c in (2,15,16):
            cell.number_format = 'MM/DD/YYYY'
            cell.alignment = CENTER

    # Bold the order ID in first occurrence
    ws_ord.cell(row=r, column=1).font = Font(name="Arial", size=10, bold=True)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — INVOICE TRACKER
# ═══════════════════════════════════════════════════════════════════════════════
ws_inv = wb.create_sheet("Invoice Tracker")
ws_inv.sheet_view.showGridLines = False
ws_inv.freeze_panes = "A3"
ws_inv.row_dimensions[1].height = 36
ws_inv.row_dimensions[2].height = 22

ws_inv.merge_cells("A1:L1")
t = ws_inv["A1"]
t.value = "MYLD — Invoice Tracker"
t.font  = TITLE_FONT; t.alignment = LEFT; t.fill = WHITE_FILL

inv_headers = ["Invoice #","Order ID","Account Name","State","Invoice Date",
               "Due Date","Subtotal","Tax Rate","Tax Amount","Total",
               "Payment Status","Days Overdue"]
inv_widths  = [14,14,26,7,14,14,14,12,14,14,16,14]
for i, (h, w) in enumerate(zip(inv_headers, inv_widths), 1):
    header_cell(ws_inv, 2, i, h, width=w)

invoices = [
    ("INV-001","ORD-A1B2C3D4","Green Leaf Dispensary","OR",
     today - datetime.timedelta(days=28), today + datetime.timedelta(days=2),
     270.00, 0.17, "Paid"),
    ("INV-002","ORD-E5F6G7H8","The Healing Center","CA",
     today - datetime.timedelta(days=12), today + datetime.timedelta(days=18),
     216.00, 0.15, "Open"),
]

PAYMENT_COLORS = {"Paid": "D5F5E3", "Open": "D6EAF8", "Overdue": "FADBD8", "Partial": "FDEBD0"}

for r, inv in enumerate(invoices, 3):
    ws_inv.row_dimensions[r].height = 18
    pay_status = inv[8]
    row_fill = PatternFill("solid", fgColor=PAYMENT_COLORS.get(pay_status, "FFFFFF"))
    sub = inv[6]; tax_rate = inv[7]
    tax_amt = f"=G{r}*H{r}"
    total   = f"=G{r}+J{r}"
    overdue = f'=IF(K{r}="Paid","",MAX(0,TODAY()-F{r}))'
    vals = [inv[0],inv[1],inv[2],inv[3],inv[4],inv[5],sub,tax_rate,tax_amt,total,pay_status,overdue]
    for c, v in enumerate(vals, 1):
        cell = ws_inv.cell(row=r, column=c, value=v)
        cell.fill = row_fill; cell.font = BODY_FONT; cell.border = full_border()
        cell.alignment = CENTER if c in (1,2,4,8,11,12) else LEFT
        if c in (7,9,10): cell.number_format = '$#,##0.00'; cell.alignment = RIGHT
        if c == 8:        cell.number_format = '0.0%'
        if c in (5,6):    cell.number_format = 'MM/DD/YYYY'; cell.alignment = CENTER
    pay_cell = ws_inv.cell(row=r, column=11)
    if pay_status == "Paid":
        pay_cell.font = Font(name="Arial", size=10, bold=True, color="2D6A4F")
    elif pay_status == "Overdue":
        pay_cell.font = Font(name="Arial", size=10, bold=True, color="C0392B")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — CREDIT MEMO LOG
# ═══════════════════════════════════════════════════════════════════════════════
ws_cm = wb.create_sheet("Credit Memo Log")
ws_cm.sheet_view.showGridLines = False
ws_cm.freeze_panes = "A3"
ws_cm.row_dimensions[1].height = 36
ws_cm.row_dimensions[2].height = 22

ws_cm.merge_cells("A1:K1")
t = ws_cm["A1"]
t.value = "MYLD — Credit Memo Log"
t.font  = TITLE_FONT; t.alignment = LEFT; t.fill = WHITE_FILL

cm_headers = ["Memo #","Invoice #","Order ID","Account Name","State","Memo Date",
              "SKU","Return Qty","Unit Price","Credit Amount","Reason","Notes"]
cm_widths  = [14,14,14,26,7,14,22,12,12,14,22,34]
for i, (h, w) in enumerate(zip(cm_headers, cm_widths), 1):
    header_cell(ws_cm, 2, i, h, width=w)

cm_data = [
    ("CM-001","INV-002","ORD-E5F6G7H8","The Healing Center","CA",
     today - datetime.timedelta(days=10),
     "WLD-GUM-HUCK-10",2,18.00,"Damaged in Transit","Product crushed in transit — 2 units unsellable"),
]

for r, row in enumerate(cm_data, 3):
    ws_cm.row_dimensions[r].height = 18
    row_fill = GRAY_FILL if r % 2 == 0 else WHITE_FILL
    qty_col  = 8
    px_col   = 9
    credit   = f"=H{r}*I{r}"
    vals = list(row[:9]) + [credit] + list(row[9:])
    for c, v in enumerate(vals, 1):
        cell = ws_cm.cell(row=r, column=c, value=v)
        cell.fill = row_fill; cell.font = BODY_FONT; cell.border = full_border()
        cell.alignment = CENTER if c in (1,2,3,5,8) else LEFT
        if c == 9:  cell.number_format = '$#,##0.00'
        if c == 10: cell.number_format = '$#,##0.00'; cell.alignment = RIGHT
        if c == 6:  cell.number_format = 'MM/DD/YYYY'; cell.alignment = CENTER


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
ws_dash = wb.create_sheet("Dashboard")
ws_dash.sheet_view.showGridLines = False

# Move Dashboard to front
wb.move_sheet("Dashboard", offset=-(len(wb.sheetnames)-1))

ws_dash.row_dimensions[1].height = 50
ws_dash.row_dimensions[2].height = 14
ws_dash.column_dimensions["A"].width = 2

# Title bar
ws_dash.merge_cells("B1:L1")
t = ws_dash["B1"]
t.value = "MYLD — Sales Operations Dashboard"
t.font  = Font(name="Arial", bold=True, size=18, color=WHITE)
t.alignment = LEFT
t.fill = GREEN_FILL

# KPI section label
ws_dash.merge_cells("B3:L3")
kpi_label = ws_dash["B3"]
kpi_label.value = "ORDER PIPELINE SUMMARY"
kpi_label.font  = Font(name="Arial", bold=True, size=10, color=DARK_GRAY)
kpi_label.alignment = LEFT

# KPI boxes
kpi_data = [
    ("B","C","Active Orders",    '=COUNTIF(\'Order Log\'!M:M,"Submitted")+COUNTIF(\'Order Log\'!M:M,"Confirmed")+COUNTIF(\'Order Log\'!M:M,"Fulfillment")'),
    ("D","E","Pending Invoice",  '=COUNTIF(\'Order Log\'!M:M,"Fulfillment")'),
    ("F","G","Invoiced (Open)",  '=COUNTIF(\'Invoice Tracker\'!K:K,"Open")'),
    ("H","I","Credit Memos",     '=COUNTA(\'Credit Memo Log\'!A3:A100)-1'),
    ("J","K","Closed This Month",'=COUNTIF(\'Order Log\'!M:M,"Closed")'),
]

ws_dash.row_dimensions[4].height = 14
ws_dash.row_dimensions[5].height = 48
ws_dash.row_dimensions[6].height = 24
ws_dash.row_dimensions[7].height = 14

for start_col, end_col, label, formula in kpi_data:
    sc = ord(start_col) - ord('A') + 1
    ec = ord(end_col)   - ord('A') + 1
    ws_dash.merge_cells(f"{start_col}5:{end_col}5")
    ws_dash.merge_cells(f"{start_col}6:{end_col}6")
    ws_dash.column_dimensions[start_col].width = 12
    ws_dash.column_dimensions[end_col].width   = 2

    box = ws_dash[f"{start_col}5"]
    box.value = formula
    box.font  = Font(name="Arial", bold=True, size=28, color=WYLD_GREEN)
    box.fill  = LIGHT_FILL
    box.alignment = CENTER
    box.border = full_border()

    lbl = ws_dash[f"{start_col}6"]
    lbl.value = label
    lbl.font  = Font(name="Arial", size=9, color=DARK_GRAY)
    lbl.fill  = ACCENT_FILL
    lbl.alignment = CENTER
    lbl.border = full_border()

    # merge the end column cells too for visual box
    ws_dash[f"{end_col}5"].fill = LIGHT_FILL; ws_dash[f"{end_col}5"].border = full_border()
    ws_dash[f"{end_col}6"].fill = ACCENT_FILL; ws_dash[f"{end_col}6"].border = full_border()

# Divider
ws_dash.row_dimensions[8].height = 10

# Order Status Breakdown table
ws_dash.merge_cells("B9:E9")
ws_dash["B9"].value = "ORDER STATUS BREAKDOWN"
ws_dash["B9"].font  = Font(name="Arial", bold=True, size=10, color=DARK_GRAY)
ws_dash.row_dimensions[9].height = 20

status_breakdown = [
    ("Draft",      '=COUNTIF(\'Order Log\'!M:M,"Draft")'),
    ("Submitted",  '=COUNTIF(\'Order Log\'!M:M,"Submitted")'),
    ("Confirmed",  '=COUNTIF(\'Order Log\'!M:M,"Confirmed")'),
    ("Fulfillment",'=COUNTIF(\'Order Log\'!M:M,"Fulfillment")'),
    ("Invoiced",   '=COUNTIF(\'Order Log\'!M:M,"Invoiced")'),
    ("Closed",     '=COUNTIF(\'Order Log\'!M:M,"Closed")'),
    ("Cancelled",  '=COUNTIF(\'Order Log\'!M:M,"Cancelled")'),
    ("Rejected",   '=COUNTIF(\'Order Log\'!M:M,"Rejected")'),
]

header_cell(ws_dash, 10, 2, "Status",  width=14)
header_cell(ws_dash, 10, 3, "Lines",   width=10)
header_cell(ws_dash, 10, 4, "% of Total", width=14)

for i, (status, formula) in enumerate(status_breakdown, 11):
    ws_dash.row_dimensions[i].height = 18
    row_fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))

    s_cell = ws_dash.cell(row=i, column=2, value=status)
    s_cell.fill = row_fill; s_cell.font = Font(name="Arial", size=10, bold=True)
    s_cell.border = full_border(); s_cell.alignment = LEFT

    n_cell = ws_dash.cell(row=i, column=3, value=formula)
    n_cell.fill = row_fill; n_cell.font = BODY_FONT
    n_cell.border = full_border(); n_cell.alignment = CENTER

    p_cell = ws_dash.cell(row=i, column=4,
        value=f"=IF(COUNTA('Order Log'!M4:M500)=0,0,C{i}/COUNTA('Order Log'!M4:M500))")
    p_cell.fill = row_fill; p_cell.font = BODY_FONT
    p_cell.number_format = '0%'; p_cell.border = full_border(); p_cell.alignment = CENTER

# Revenue summary
ws_dash.merge_cells("G9:K9")
ws_dash["G9"].value = "REVENUE SUMMARY"
ws_dash["G9"].font  = Font(name="Arial", bold=True, size=10, color=DARK_GRAY)

header_cell(ws_dash, 10, 7, "Metric", width=24)
header_cell(ws_dash, 10, 8, "Amount ($)", width=16)

rev_rows = [
    ("Total Order Value (All)",     "=SUMIF('Order Log'!C:C,\"ACC*\",'Order Log'!J:J)"),
    ("Total Fulfilled Value",        "=SUMIF('Order Log'!C:C,\"ACC*\",'Order Log'!L:L)"),
    ("Total Invoiced (Open)",        "=SUMIF('Invoice Tracker'!K:K,\"Open\",'Invoice Tracker'!J:J)"),
    ("Total Credit Memos Issued",    "=SUMIF('Credit Memo Log'!A:A,\"CM*\",'Credit Memo Log'!J:J)"),
]

for i, (label, formula) in enumerate(rev_rows, 11):
    ws_dash.row_dimensions[i].height = 18
    row_fill = GRAY_FILL if i % 2 == 0 else WHITE_FILL
    l = ws_dash.cell(row=i, column=7, value=label)
    l.fill = row_fill; l.font = BODY_FONT; l.border = full_border(); l.alignment = LEFT
    v = ws_dash.cell(row=i, column=8, value=formula)
    v.fill = row_fill; v.font = Font(name="Arial", size=10, bold=True, color=WYLD_GREEN)
    v.number_format = '$#,##0.00'; v.border = full_border(); v.alignment = RIGHT

# Quick links section
ws_dash.row_dimensions[20].height = 14
ws_dash.merge_cells("B21:L21")
ws_dash["B21"].value = "QUICK REFERENCE — State Return Rules"
ws_dash["B21"].font  = Font(name="Arial", bold=True, size=10, color=DARK_GRAY)

header_cell(ws_dash, 22, 2, "State")
header_cell(ws_dash, 22, 3, "Returns?", width=12)
header_cell(ws_dash, 22, 4, "Window", width=12)
header_cell(ws_dash, 22, 5, "Tax Rate", width=12)

quick_states = [
    ("OR","No","—","17%"),("WA","No","—","37%"),("CA","Yes","30 days","15%"),
    ("CO","Yes","14 days","15%"),("NV","No","—","10%"),("MI","Yes","30 days","10%"),
    ("IL","No","—","10%"),("NJ","No","—","7%"),("NY","Yes","14 days","9%"),("MA","Yes","30 days","10.7%"),
]
for i, (st, ret, win, tax) in enumerate(quick_states, 23):
    ws_dash.row_dimensions[i].height = 16
    row_fill = GRAY_FILL if i % 2 == 0 else WHITE_FILL
    for c, v in enumerate([st, ret, win, tax], 2):
        cell = ws_dash.cell(row=i, column=c, value=v)
        cell.fill = row_fill; cell.font = BODY_FONT
        cell.border = full_border(); cell.alignment = CENTER
    ret_cell = ws_dash.cell(row=i, column=3)
    if ret == "Yes":
        ret_cell.font = Font(name="Arial", size=10, bold=True, color="2D6A4F")
    else:
        ret_cell.font = Font(name="Arial", size=10, bold=True, color="C0392B")

